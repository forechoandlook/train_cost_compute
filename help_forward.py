
# how to generate forward graph and make it into table format
import json 
from collections import defaultdict
import math
import numpy as np
from ops import *
import pandas as pd
from itertools import product
from copy import deepcopy
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

class mylog:
    @staticmethod
    def info(*args, **kwargs):
        print(*args, **kwargs)

path = "/Users/wangyangzuo/Desktop/公司/sd_forward.json"

total_graph = json.load(open(path,'r'))

graph         = total_graph['graph']
links         = total_graph['links']
reverse_links = defaultdict(list)
in_degree     = defaultdict(int)

for k,v in links.items():
    for link in v:
        reverse_links[link].append(k)
        in_degree[k] += 1

def find_grad_nodes(graph):
    grad_node_ids   = []
    grad_module_ids = []
    next_nodes = [[graph, None]]
    while next_nodes:
        cur_node, father_node = next_nodes.pop()
        k = list(cur_node.keys())[0]
        father_k = list(father_node.keys())[0] if father_node else None
        if father_k and father_node[father_k]['need_train']:
            grad_node_ids.append(k)
            grad_module_ids.append([father_k,father_node[father_k]["name"], father_node[father_k]["input_dtype"][0],father_node[father_k]["comment"]])
        if not cur_node[k]["children"]:
            continue
        for child in cur_node[k]['children'][::-1]:
            next_nodes.append([child, cur_node])
    return grad_node_ids, grad_module_ids

mylog.info(">>>>> find grad nodes")

grad_node_ids,grad_module = find_grad_nodes(graph)

def get_linear_grad_info(single_module):
    dtype = single_module[2]
    info  = single_module[3]
    in_features  = int(info.split("in_features=")[1].split(",")[0])
    out_features = int(info.split("out_features=")[1].split(",")[0])
    bias = info.split("bias=")[1].split(",")[0] == "True"  
    mem_info = in_features * out_features * dtype_map[dtype] + (out_features if bias else 0) * dtype_map[dtype]
    return mem_info, [out_features, in_features], dtype

def parse_grad_module_info(grad_module):
    grads = []
    grad_mem = 0
    adam_mem = 0
    for each_grad in grad_module:
        if each_grad[1] == "Linear":
            mem_info, shape, dtype = get_linear_grad_info(each_grad)
            grad_mem += mem_info
            adam_mem += mem_info * 2 if each_grad[2] == "float32" else mem_info * 4
            grads.append([shape, dtype])
        # grads.append(each_grad)
    return grads, grad_mem, adam_mem

grads, grad_mem, adam_mem = parse_grad_module_info(grad_module)



import pdb;pdb.set_trace()
mylog.info(">>>>> find all activation node")
starts = [i for i in grad_node_ids]
activation_nodes = set()
while starts:
    start = starts.pop()
    activation_nodes.add(start)
    if start not in links:
        continue
    for link in links[start]:
        if str(link) not in activation_nodes:
            starts.append(str(link))

mylog.info(">>>>> find all loss nodes")
# find all loss nodes
starts = [i for i in grad_node_ids]
loss_nodes =[]
vis = set()
while starts:
    start = starts.pop()
    vis.add(start)
    if str(start) not in links:
        loss_nodes.append(start)
        continue
    for link in links[str(start)]:
        if link not in vis:
            starts.append(link)

mylog.info(">>>>> match shape")

model_param = total_graph.get('model_param', {
    "n"     : 1,
    "c"     : 4,
    "h"     : 64,
    "w"     : 64,
    "seq"   : 77,
    "dtype" : "float16",
    "rank"  : 4
})

hw_region  = [64*64, 64*64/4, 64*64/16, 64*64/64]
h_w_region = [64, 32, 16, 8]

class layout:
    # 4 dim
    n_model_h_w       = "n_model_h_w"
    n_h_w_model       = "n_h_w_model"
    n_model_hw_model  = "n_model_hw_model"
    n_hw_model_model  = "n_hw_model_model"
    n_seq_model_model = "n_seq_model_model"
    n_model_model_model = "n_model_model_model"
    n_model_seq_model = "n_model_seq_model"
    # 3 dim
    n_seq_model       = "n_seq_model"
    n_seq_r           = "n_seq_r"
    nmodel_model_hw   = "nmodel_model_hw"
    nmodel_hw_seq     = "nmodel_hw_seq"
    nmodel_model_seq  = "nmodel_model_seq"
    nmodel_seq_model  = "nmodel_seq_model"
    n_hw_r            = "n_hw_r"
    # 3 dim
    n_hw_model        = "n_hw_model"
    n_hw_hw           = "n_hw_hw"
    nmodel_hw_hw      = "nmodel_hw_hw"
    nmodel_hw_model   = "nmodel_hw_model"
    # 2 dim
    n_model           = "n_model"
    # 1 dim
    n                 = "n"
    model             = "model"
 
    @staticmethod
    def match_shape(shape: list):
        if len(shape) == 1:
            if shape[0] == model_param['n']:
                return layout.n, [1]
            return layout.model, [0]
        if len(shape) == 2:
            return layout.n_model, [1,0]
        if len(shape) == 3:
            if shape[0] == model_param['n']:
                if shape[1] == model_param["seq"]:
                    if shape[2] == model_param["rank"]:
                        return layout.n_seq_r, [1,shape[1] / model_param["seq"], shape[2] / model_param["rank"]]
                    else:
                        return layout.n_seq_model, [1, shape[1] / model_param["seq"], 0]
                elif shape[2] == model_param["rank"]:
                    return layout.n_hw_r, [1, shape[1] / (model_param["h"] * model_param["w"]), shape[2] / model_param["rank"]]
                elif shape[1] == shape[2]:
                    return layout.n_hw_hw, [1, shape[1] / (model_param["h"] * model_param["w"]), shape[2] / (model_param["h"] * model_param["w"])]
                else:
                    return layout.n_hw_model, [1, shape[1] / (model_param["h"] * model_param["w"]), 0]
            # 3 dim not starts with n
            elif shape[1] == model_param["seq"]:
                return layout.nmodel_seq_model, [1, shape[1] / model_param["seq"], 0]
            elif shape[2] == model_param["seq"] and shape[1] in h_w_region:
                # nmodel_hw_seq
                return layout.nmodel_hw_seq, [1, shape[1] / (model_param["h"] * model_param["w"]), shape[2] / model_param["seq"]]
            elif shape[2] == model_param["seq"]:
                # nmodel_model_seq
                return layout.nmodel_model_seq, [1, 0, shape[2] / model_param["seq"]]
            elif shape[1] == shape[2] and shape[1] in hw_region:
                # nmodel_hw_hw
                return layout.nmodel_hw_hw, [1, shape[1] / (model_param["h"] * model_param["w"]), shape[2] / (model_param["h"] * model_param["w"])]
            elif shape[2] in hw_region:
                return layout.nmodel_model_hw, [1, 0, shape[2] / (model_param["h"] * model_param["w"])]
            else:
                return layout.nmodel_hw_model, [1, shape[1] / (model_param["h"] * model_param["w"]), 0]
        if len(shape) == 4:
            if shape[0] == model_param['n']:
                if shape[1] == model_param["seq"]:
                    return layout.n_seq_model_model, [1, shape[1] / model_param["seq"], 0, 0]
                elif shape[1] in h_w_region and shape[2] in h_w_region and shape[1] == shape[2]:
                    return layout.n_h_w_model, [1, shape[1] / model_param["h"], shape[2] / model_param["w"], 0]
                elif shape[2] in h_w_region and shape[3] in h_w_region:
                    return layout.n_model_h_w, [1, 0, shape[2] / model_param["h"], shape[3] / model_param["w"]]
                elif shape[1] in hw_region:
                    return layout.n_hw_model_model, [1, shape[1] / (model_param["h"] * model_param["w"]), 0, 0]
                elif shape[2] in hw_region:
                    return layout.n_model_hw_model, [1, 0, shape[2] / (model_param["h"] * model_param["w"]), 0]
                elif shape[2] == model_param["seq"]:
                    return layout.n_model_seq_model, [1, 0, shape[2] / model_param["seq"], 0]
                else:
                    return layout.n_model_model_model, [1, 0, 0, 0]
            else:
                import pdb;pdb.set_trace()
                print("shape error")
        return 0,[0]

    @staticmethod
    def match_layout(node):
        input_shapes = node["input_shape"]
        output_shapes = node["output_shape"]
        node["input_shape_layout"]  = [layout.match_shape(i) for i in input_shapes]
        node["output_shape_layout"] = [layout.match_shape(i) for i in output_shapes]
        return node

    @staticmethod
    def match_layout_source_node(node):
        k = list(node.keys())[0]
        input_shapes = node[k]["input_shape"]
        output_shapes = node[k]["output_shape"]
        node[k]["input_shape_layout"]  = [layout.match_shape(i) for i in input_shapes]
        node[k]["output_shape_layout"] = [layout.match_shape(i) for i in output_shapes]
        return node

    @staticmethod
    def fix_single_shape(shape, cur_layout, rate, param):
        # param 修复过的 only support nhw
        if cur_layout == layout.n:
            return [param["n"]]
        if cur_layout == layout.model:
            return shape
        if cur_layout == layout.n_model:
            return [param["n"], shape[1]]
        if cur_layout == layout.n_seq_r:
            return [param["n"], shape[1], shape[2]]
        if cur_layout == layout.n_seq_model:
            return [param["n"], shape[1], shape[2]]
        if cur_layout == layout.n_hw_r:
            return [param["n"], param["h"] * param["w"] * rate[1], shape[2]]
        if cur_layout == layout.n_hw_hw:
            return [param["n"], param["h"] * param["w"] * rate[1], param["h"] * param["w"] * rate[2]]
        if cur_layout == layout.n_hw_model:
            return [param["n"], param["h"] * param["w"] * rate[1], shape[2]]
        if cur_layout == layout.nmodel_seq_model:
            return [shape[0] * param["n"], shape[1], shape[2]]
        if cur_layout == layout.nmodel_hw_seq:
            return [shape[0] * param["n"], param["h"] * param["w"] * rate[1], shape[2]]
        if cur_layout == layout.nmodel_model_seq:
            return [shape[0] * param["n"], shape[1], shape[2]]
        if cur_layout == layout.nmodel_hw_hw:
            return [shape[0] * param["n"], param["h"] * param["w"] * rate[1], param["h"] * param["w"] * rate[2]]
        if cur_layout == layout.nmodel_hw_model:
            return [shape[0] * param["n"], param["h"] * param["w"] * rate[1], shape[2]]
        if cur_layout == layout.nmodel_model_hw:
            return [shape[0] * param["n"], shape[1], param["h"] * param["w"] * rate[2]]
        if cur_layout == layout.n_seq_model_model:
            return [param["n"], shape[1], shape[2], shape[3]]
        if cur_layout == layout.n_h_w_model:
            return [param["n"], param["h"] * rate[1], param["w"] * rate[2], shape[3]]
        if cur_layout == layout.n_model_h_w:
            return [param["n"], shape[1], param["h"] * rate[2], param["w"] * rate[3]]
        if cur_layout == layout.n_model_model_model:
            return [param["n"], shape[1], shape[2], shape[3]]
        if cur_layout == layout.n_hw_model_model:
            return [param["n"], param["h"] * param["w"] * rate[1], shape[2], shape[3]]
        if cur_layout == layout.n_model_hw_model:
            return [param["n"], shape[1], param["h"] * param["w"] * rate[2], shape[3]]
        if cur_layout == layout.n_model_seq_model:
            return [param["n"], shape[1], shape[2], shape[3]]
        print("error: not support layout")
        import pdb;pdb.set_trace()
    
    @staticmethod
    def fix_shape(node, param):
        # param: {n,h,w,r,seq}
        k                = list(node.keys())[0]
        input_shapes     = node[k]["input_shape"]
        output_shapes    = node[k]["output_shape"]
        input_layer_out  = node[k]["input_shape_layout"]
        output_layer_out = node[k]["output_shape_layout"]
        new_input_shape  = []
        new_output_shape = []
        # print("fix shape", input_shapes, output_shapes)
        # print("layout ", input_layer_out, output_layer_out)
        for i in range(len(input_shapes)):
            cur_layout, rate = input_layer_out[i]
            t = layout.fix_single_shape(input_shapes[i], cur_layout, rate, param)
            new_input_shape.append([int(i) for i in t])
        for i in range(len(output_shapes)):
            cur_layout, rate = output_layer_out[i]
            t = layout.fix_single_shape(output_shapes[i], cur_layout, rate, param)
            new_output_shape.append([int(i) for i in t])
        # print("after",  new_input_shape, new_output_shape)
        # print(">>>>>>>>>>>>> ", node[k]["name"], k, " <<<<<<<<<<<<<<<<")
        node[k]["input_shape"] = new_input_shape
        node[k]["output_shape"]= new_output_shape
        return node

# init: 输入的权重，梯度权重 adam 
weights         = total_graph.get('weights', 2041.164e6 )
grad_weights    = grad_mem
adam_weights    = adam_mem

all_grad_shapes = grads
first_key = list(graph.keys())[0]
start_depth = 0

def build_preview_mem():
    
    pass


def match_shape():
    next_nodes = [[graph, None]]
    vis = set()
    while next_nodes:
        cur_node, father_node = next_nodes.pop()
        k = list(cur_node.keys())[0]
        if k in vis: continue
        if not cur_node[k]["children"]:
            layout.match_layout_source_node(cur_node)
            vis.add(k)
            continue
        else:
            layout.match_layout_source_node(cur_node)
        vis.add(k)
        for child in cur_node[k]['children'][::-1]:
            next_nodes.append([child, cur_node])
mylog.info(">>>>> match shape")
match_shape()
source_graph = deepcopy(graph)



# fix shape 


def fix_shape_total_graph(cur_shapes):
    next_nodes = [[graph, None]]
    vis = set()
    while next_nodes:
        cur_node, father_node = next_nodes.pop()
        k = list(cur_node.keys())[0]
        if k in vis: continue
        if not cur_node[k]["children"]:
            layout.fix_shape(cur_node, cur_shapes)
            continue
        else:
            layout.fix_shape(cur_node, cur_shapes)
        for child in cur_node[k]['children'][::-1]:
            next_nodes.append([child, cur_node])

def get_output_mem(node):
    k = list(node.keys())[0]
    name = node[k]['name']
    if name in ["reshape", "__getitem__", "to","chunk", "dropout", "contiguous", "float", "permute", "transpose"]:
        return 0
    output_shapes = node[k]['output_shape']
    output_dtype = node[k]['output_dtype']
    output_dtypes = [ dtype_map[i] for i in output_dtype]
    out_mem = sum([np.prod(i) * j for i, j in zip(output_shapes, output_dtypes)])
    return int(out_mem)

def get_output_mem_node(node):
    k = node["id"]
    name = node['name']
    if name in ["reshape", "__getitem__", "to","chunk", "contiguous", "float", "permute", "transpose"]:
        return 0
    output_shapes = node['output_shape']
    output_dtype = node['output_dtype']
    output_dtypes = [ dtype_map[i] for i in output_dtype]
    out_mem = sum([np.prod(i) * j for i, j in zip(output_shapes, output_dtypes)])
    return int(out_mem)

def activation_has_outputs_check_node(node):
    k = node["id"]
    name = node["name"]
    if name in ["baddbmm", "mul", "add", "cat"]:
        return False
    if k in activation_nodes:
        return True
    return False

def activation_has_outputs_node(node):
    k = list(node.keys())[0]
    name = node[k]['name']
    if name in ["baddbmm", "mul", "add"]:
        return False
    if k in activation_nodes:
        return True
    return False

def walk_for_training_mem_activation_mem(graph):
    # grad_node_ids = []
    next_nodes = [[graph, None]]
    module_stack = []
    module_stack_node = []
    idx = 0
    activation_mem_usage = defaultdict(int)
    while next_nodes:
        cur_node, father_node = next_nodes.pop()
        k = list(cur_node.keys())[0]
        father_k = list(father_node.keys())[0] if father_node else None
        while father_k and module_stack and father_k != module_stack[-1]:
            t_k = module_stack.pop()
            t   = module_stack_node.pop()
            cur_t_k = module_stack[-1] if module_stack else None
            cur_t   = module_stack_node[-1] if module_stack else None
            if cur_t_k:
                cur_t[cur_t_k]["module_mem"] += get_output_mem(t)
                cur_t[cur_t_k]["graph_mem"]  += t[t_k]["graph_mem"]
                cur_t[cur_t_k]["cur_mem"] = max(cur_t[cur_t_k]["module_mem"], cur_t[cur_t_k]["graph_mem"] )
                t[t_k]["max_mem"]         = max(t[t_k]["max_mem"], t[t_k]["module_mem"])
                cur_t[cur_t_k]["max_mem"] = max(cur_t[cur_t_k]["max_mem"], t[t_k]["max_mem"], cur_t[cur_t_k]["module_mem"] )
        
        if cur_node[k]["children"]:
            module_stack.append(k)
            module_stack_node.append(cur_node)
            cur_node[k]["module_mem"] = 0
            cur_node[k]["graph_mem"]  = 0
            cur_node[k]["max_mem"]    = 0
            cur_node[k]["cur_mem"]    = max(father_node[father_k]["module_mem"], father_node[father_k]["graph_mem"] )if father_k else 0
        
        if not cur_node[k]['children']:
            cur_node[k]["idx"] = idx
            idx += 1
            cur_mem = get_output_mem(cur_node)
            cur_node[k]['cur_mem'] = father_node[father_k]["cur_mem"] + cur_mem
            cur_node[k]["module_mem"] = cur_mem

            if activation_has_outputs_node(cur_node):
                cur_node[k]["graph_mem"] = cur_mem

            father_node[father_k]["module_mem"] += cur_mem
            father_node[father_k]["cur_mem"]    += cur_mem
            if activation_has_outputs_node(cur_node):
                activation_mem_usage[cur_node[k]["name"]] += cur_mem
                father_node[father_k]["graph_mem"] += cur_mem
            continue

        for child in cur_node[k]['children'][::-1]:
            next_nodes.append([child, cur_node])
    return activation_mem_usage

def forward_dma_ops_calc():
    res_set = {}
    def handle_record_node(node, pre):
        node_key                 = list(node.keys())[0]
        pre_key                  = list(pre.keys())[0]
        res_node                 = {}
        depth                    = node[node_key]['depth']
        node[node_key]['info']   = pre[pre_key]["comment"]
        if node[node_key]["name"] == "empty-pass":
            return 
        if node[node_key]["name"] in fix_op:
            node[node_key] = fix_op[node[node_key]["name"]](node[node_key])
        # layout.match_layout(node[node_key])
        node[node_key]           = ops_info[node[node_key]["name"]](node[node_key])
        res_node["name"]         = node[node_key]["name"]
        res_node['id']           = node_key
        res_node['depth']        = node[node_key]['depth']
        node[node_key]['info']   = pre[pre_key]["comment"][:200] if depth > 1 else ""
        res_node['path']         = node[node_key]['path']
        res_node["info"]         = pre[pre_key]["comment"][:200] if depth > 1 else ""
        res_node['align_input_shape']  = node_fn[res_node['name']](node)
        res_node["input_shape"]        = node[node_key]['input_shape']
        res_node["output_shape"]       = node[node_key]["output_shape"]
        res_node['align_output_shape'] = [ handle_input_shape(node[node_key]['output_shape'][0])]
        res_node['align_input_dtype']  = node[node_key]['input_dtype'][0]
        res_node['align_output_dtype'] = node[node_key]['output_dtype'][0]
        res_node['input_dtype']        = node[node_key]['input_dtype']
        res_node['output_dtype']       = node[node_key]['output_dtype']
        res_node["ops"]                = node[node_key]["ops"]
        res_node["s2l_dma"]            = node[node_key]["s2l_dma"]
        res_node["l2s_dma"]            = node[node_key]["l2s_dma"]
        res_node["input_shape_layout"] = node[node_key]["input_shape_layout"]
        res_node["output_shape_layout"] = node[node_key]["output_shape_layout"]
        res_node["cur_mem"]      = node[node_key]["cur_mem"]
        res_node["forward_dma"]  = res_node["s2l_dma"] + res_node["l2s_dma"]
        if res_node["name"] == "conv2d":
            res_node["kernel_shape"] = res_node["align_input_shape"][1]
        # res_set.append(res_node)
        res_set[node_key] = res_node
        return res_node
    next_nodes = [[graph, None]]
    need_mem = defaultdict(dict)
    tables = []
    while next_nodes:
        cur_node, father_node = next_nodes.pop()
        k = list(cur_node.keys())[0]
        father_k = list(father_node.keys())[0] if father_node else None
        if cur_node[k]['depth'] >= 0:
            if "path" not in father_node[father_k]:
                cur_node[k]["path"] = [ cur_node[k]["name"] ]
            else:
                cur_node[k]["path"] = father_node[father_k]["path"] + [ cur_node[k]["name"] ]
        if not cur_node[k]["children"]:
            handle_record_node(cur_node, father_node)
            continue
        for child in cur_node[k]['children'][::-1]:
            next_nodes.append([child, cur_node])
    return res_set

# get back_grad_idx
def insert_back_grad_idx(node):
    node_id = node['id']
    if node_id not in activation_nodes: return node
    name = node['name']
    if node_id in grad_node_ids:
        node["leaf"] = True
    else:
        node["leaf"] = False
    if node["leaf"]:
        if name == "conv":
            # check bais 
            has_bias = node.get("bias", False)
            node["back_grad_idx"] = list(range(3)) if has_bias else list(range(2))
        elif name == "linear":
            # check bias
            has_bias = node.get("bias", False)
            node["back_grad_idx"] = list(range(3)) if has_bias else list(range(2))
        elif name == "layer_norm":
            # check affine
            node["back_grad_idx"] = list(range(3))
        elif name == "group_norm":
            # check affine
            node["back_grad_idx"] = list(range(3))
        else:
            print("error: not support leaf node")
            exit(1)
    else:
        # 看输入是否是激活节点，如果是就算
        # 有一些只有一个输入或者是输入较少，比如说mul 构造隐藏的tensor节点
        input_ids = [i for i in reverse_links[int(node_id)] if i in activation_nodes]
        # match inputs node
        input_num = len(node["input_shape"])
        if name in ["mul", "add", "div", "sub"] and input_num == 1:
            node["back_grad_idx"] = [0]
        if input_num == len(input_ids):
            node["back_grad_idx"] = list(range(input_num))
        else:
            input_ids_shapes = [res_set[i]["output_shape"][0] for i in input_ids]# 万一是多输出咋整？（特例再说吧） TODO fix it
            res = []
            for i in range(input_num):
                if node["input_shape"][i] in input_ids_shapes:
                    res.append(i)
            node["back_grad_idx"] = res
    return node

def prepare_bwd_node_outputs_calc(res_set):
    for k in res_set.keys():
        res_set[k] = insert_back_grad_idx(res_set[k])
    return res_set

def calc_bwd_bdc_dma(res_set):
    backward_start = [ str(i) for i in loss_nodes]
    print("backward start",backward_start)
    grad_set = set(grad_node_ids)
    cur_back_in_degress = deepcopy(in_degree)
    while backward_start:
        start = backward_start.pop(0)
        if start in activation_nodes and start in res_set:
            # calculate 
            # backward_node_fn
            node = res_set[start]
            name = node["name"]
            calc_dma_backops(node)
            node = backward_node_fn[name](node)
            # remove mem usage
            if activation_has_outputs_check_node(node):
                node["back_mem"] = node["cur_mem"] - get_output_mem_node(node)
        if start in grad_set:
            grad_set.remove(start)
        if len(grad_set) == 0:
            print("have found all the backward nodes")
            break
        for v in reverse_links[int(start)]:
            cur_back_in_degress[v] -= 1
            if cur_back_in_degress[v] == 0:
                backward_start.append(v)
    return res_set

def calc_all_time(res_set):
    total_times = {
        "1684x_forward": 0,
        "2260_forward": 0,
        "1684x_backward": 0,
        "2260_backward": 0,
        "1684x_fwd_dma": 0,
        "2260_fwd_dma": 0,
        "1684x_fwd_tiu": 0,
        "2260_fwd_tiu": 0,
        "1684x_bwd_dma": 0,
        "2260_bwd_dma": 0,
        "1684x_bwd_tiu": 0,
        "2260_bwd_tiu": 0
    }
    ops_total_times = defaultdict(dict)
    for k in res_set.keys():
        # forward time 
        name = res_set[k]["name"]
        res_set[k]["1684x_forward_dma_time"] = res_set[k]["forward_dma"] / params["1684x dma"]
        res_set[k]["2260_forward_dma_time"]  = res_set[k]["forward_dma"] / params["2260 dma"]
        res_set[k]["1684x_forward_ops_time"] = res_set[k]["ops"] / (params["1684x f16 tiu"] if res_set[k]["input_dtype"][0] == "float16" else params["1684x f32 tiu"])
        res_set[k]["2260_forward_ops_time"]  = res_set[k]["ops"] / (params["2260 f16 tiu"]  if res_set[k]["input_dtype"][0] == "float16" else params["2260 f32 tiu"])
        res_set[k]["1684x_forward_time"]     = max(res_set[k]["1684x_forward_dma_time"],res_set[k]["1684x_forward_ops_time"])
        res_set[k]["2260_forward_time"]      = max(res_set[k]["2260_forward_dma_time"],res_set[k]["2260_forward_ops_time"])
        total_times["1684x_forward"]         += res_set[k]["1684x_forward_time"]
        total_times["2260_forward"]          += res_set[k]["2260_forward_time"]
        total_times["1684x_fwd_dma"]         += res_set[k]["1684x_forward_dma_time"]
        total_times["2260_fwd_dma"]          += res_set[k]["2260_forward_dma_time"]
        total_times["1684x_fwd_tiu"]         += res_set[k]["1684x_forward_ops_time"]
        total_times["2260_fwd_tiu"]          += res_set[k]["2260_forward_ops_time"]
        if "1684x_forward" not in ops_total_times[name]:
            ops_total_times[name]["1684x_forward"] = 0
            ops_total_times[name]["2260_forward"]  = 0
            ops_total_times[name]["1684x_dma"]     = 0
            ops_total_times[name]["2260_dma"]      = 0
            ops_total_times[name]["1684x_tiu"]     = 0
            ops_total_times[name]["2260_tiu"]      = 0
        ops_total_times[name]["1684x_forward"] += res_set[k]["1684x_forward_time"]
        ops_total_times[name]["2260_forward"]  += res_set[k]["2260_forward_time"]
        ops_total_times[name]["1684x_dma"]     += res_set[k]["1684x_forward_dma_time"]
        ops_total_times[name]["2260_dma"]      += res_set[k]["2260_forward_dma_time"]
        ops_total_times[name]["1684x_tiu"]     += res_set[k]["1684x_forward_ops_time"]
        ops_total_times[name]["2260_tiu"]      += res_set[k]["2260_forward_ops_time"]
        # backward time
        if k in activation_nodes:
            res_set[k]["1684x_backward_dma_time"] = res_set[k]["back_dma"] / params["1684x dma"]
            res_set[k]["2260_backward_dma_time"]  = res_set[k]["back_dma"] / params["2260 dma"]
            res_set[k]["1684x_backward_ops_time"] = res_set[k]["back_ops"] / (params["1684x f16 tiu"] if res_set[k]["input_dtype"][0] == "float16" else params["1684x f32 tiu"])
            res_set[k]["2260_backward_ops_time"]  = res_set[k]["back_ops"] / (params["2260 f16 tiu"]  if res_set[k]["input_dtype"][0] == "float16" else params["2260 f32 tiu"])
            res_set[k]["1684x_backward_time"]     = max(res_set[k]["1684x_backward_dma_time"],res_set[k]["1684x_backward_ops_time"])
            res_set[k]["2260_backward_time"]      = max(res_set[k]["2260_backward_dma_time"],res_set[k]["2260_backward_ops_time"])
            total_times["1684x_backward"]         += res_set[k]["1684x_backward_time"]
            total_times["2260_backward"]          += res_set[k]["2260_backward_time"]  
            total_times["1684x_bwd_dma"]          += res_set[k]["1684x_backward_dma_time"]
            total_times["2260_bwd_dma"]           += res_set[k]["2260_backward_dma_time"]
            total_times["1684x_bwd_tiu"]          += res_set[k]["1684x_backward_ops_time"]
            total_times["2260_bwd_tiu"]           += res_set[k]["2260_backward_ops_time"]
            if "1684x_backward" not in ops_total_times["back_"+name]:
                ops_total_times["back_"+name]["1684x_backward"] = 0
                ops_total_times["back_"+name]["2260_backward"]  = 0
                ops_total_times["back_"+name]["1684x_dma"]      = 0
                ops_total_times["back_"+name]["2260_dma"]       = 0
                ops_total_times["back_"+name]["1684x_tiu"]      = 0
                ops_total_times["back_"+name]["2260_tiu"]       = 0
            ops_total_times["back_"+name]["1684x_backward"] += max(res_set[k]["1684x_backward_ops_time"], res_set[k]["1684x_backward_dma_time"])
            ops_total_times["back_"+name]["2260_backward"]  += max(res_set[k]["2260_backward_ops_time"], res_set[k]["2260_backward_dma_time"])
            ops_total_times["back_"+name]["1684x_dma"]      += res_set[k]["1684x_backward_dma_time"]
            ops_total_times["back_"+name]["2260_dma"]       += res_set[k]["2260_backward_dma_time"]
            ops_total_times["back_"+name]["1684x_tiu"]      += res_set[k]["1684x_backward_ops_time"]
            ops_total_times["back_"+name]["2260_tiu"]       += res_set[k]["2260_backward_ops_time"]

    return res_set, total_times, ops_total_times

def write_to_sheet(ws, data, start_row, start_col):
    if isinstance(data, pd.DataFrame):
        for r_idx, row in enumerate(dataframe_to_rows(data, index=True, header=True), start=start_row):
            for c_idx, value in enumerate(row, start=start_col):
                ws.cell(row=r_idx, column=c_idx, value=value)
        return r_idx  
    elif isinstance(data, list):
        for r_idx, line in enumerate(data, start=start_row):
            ws.cell(row=r_idx, column=start_col, value=line)
        return r_idx  
    else:
        raise ValueError("Unsupported data type")

params = {
    "1684x dma": 6e4,
    "2260 dma": 548e3,
    "1684x f16 tiu": 16e6,
    "1684x f32 tiu": 2e6,
    "1684x cuba tiu": 16e6,
    "1684x vector tiu": 4e6,
    "2260 f16 tiu": 128e6,
    "2260 f32 tiu": 16e6,
}

batchs = range(1,2)
whr = [512,768,960,1024]

summary_table = defaultdict(dict)

for nb, nw in product(batchs, whr):
    nh = nw
    mylog.info("batch: %d, h: %d, w: %d"%(nb, nh, nw))
    cur_shapes = {
        "n": nb,
        "h": nh//8,
        "w": nw//8
    }
    shape_tuple = (nb, nh, nw)
    mylog.info("cur_shapes", cur_shapes)
    graph = deepcopy(source_graph)
    mylog.info(">>>>> walk for all activation mem usage")
    fix_shape_total_graph(cur_shapes)
    need_batch  = 16
    need_shapes = [512,768,960,1024]
    activation_mem_usage = walk_for_training_mem_activation_mem(graph)
    mylog.info(">>>>> make mem table for activation")
    # mem_table = pd.DataFrame.from_dict(activation_mem_usage, orient='index', columns=['op'])
    model_id = list(graph.keys())[0]
    model_graph = graph[model_id]["children"][0]
    model_graph_id = list(model_graph.keys())[0]
    graph_mem  = model_graph[model_graph_id]["graph_mem"]
    # mem_table["rate"] = mem_table["op"] / graph_mem
    # add total mem usage
    activation_mem_usage["total"] = sum(activation_mem_usage.values())
    mylog.info(">>>>> mem over")
    # forward and backward graph dma and ops calc
    res_set = forward_dma_ops_calc()
    res_set = prepare_bwd_node_outputs_calc(res_set)
    res_set = calc_bwd_bdc_dma(res_set)
    # calc backward graph and memory usage(time)
    res_set, total_times, ops_times = calc_all_time(res_set)
    fwd_bwd_df = pd.DataFrame.from_dict(res_set, orient='index')
    # path -> padding into 6
    # in table add path_0, path_1, path_2, path_3, path_4, path_5
    # fwd_bwd_df["path_0"] = fwd_bwd_df["path"].apply(lambda x: x[0] if len(x) > 0 else "")
    fwd_bwd_df["path_1"] = fwd_bwd_df["path"].apply(lambda x: x[1] if len(x) > 1 else "")
    fwd_bwd_df["path_2"] = fwd_bwd_df["path"].apply(lambda x: x[2] if len(x) > 2 else "")
    fwd_bwd_df["path_3"] = fwd_bwd_df["path"].apply(lambda x: x[3] if len(x) > 3 else "")
    fwd_bwd_df["path_4"] = fwd_bwd_df["path"].apply(lambda x: x[4] if len(x) > 4 else "")
    # fwd_bwd_df["path_5"] = fwd_bwd_df["path"].apply(lambda x: x[5] if len(x) > 5 else "")
    remove_column_names = ["id", "path", "s2l_dma", "l2s_dma", "depth", "info","depth", "info","align_input_shape", "align_output_shape", "align_input_dtype", "align_output_dtype", "input_shape_layout", "output_shape_layout", "input_dtype", "output_dtype"]
    fwd_bwd_df = fwd_bwd_df.drop(remove_column_names, axis=1)
    # reorder columns
    new_names = ["path_1", "path_2", "path_3", "path_4", "name", "forward_dma", "ops", "1684x_forward_dma_time", "1684x_forward_ops_time", "1684x_forward_time", "2260_forward_dma_time", "2260_forward_ops_time", "2260_forward_time", "back_dma", "back_ops", "1684x_backward_dma_time", "1684x_backward_ops_time", "1684x_backward_time", "2260_backward_dma_time", "2260_backward_ops_time", "2260_backward_time", "leaf",  "cur_mem",  "back_mem"]
    fwd_bwd_df = fwd_bwd_df[new_names]
    # 去掉index
    fwd_bwd_df.reset_index(drop=True, inplace=True)
    fwd_bwd_df     = fwd_bwd_df.fillna(0)
    ops_times_df   = pd.DataFrame.from_dict(ops_times, orient='index').fillna(0)
    # add total time
    ops_times_df.loc["total"] = ops_times_df.sum()
    total_times["1684x_total_time"]           = total_times["1684x_forward"] + total_times["1684x_backward"]
    total_times["2260_total_time"]            = total_times["2260_forward"]  + total_times["2260_backward"]
    total_times["total_activation"]           = graph_mem
    summary_table[shape_tuple]["total"]       = total_times
    summary_table[shape_tuple]["ops_times"]   = ops_times_df
    summary_table[shape_tuple]["mem_table"]   = activation_mem_usage
    summary_table[shape_tuple]["fwd_bwd_df"]  = fwd_bwd_df

# summary tables
pre_table = {}
# total time summary table
all_shape_total_table = {}
all_shape_mem_table   = {}
# all_shape_ops_table   = {}
for k in summary_table.keys():
    kname = "_".join(str(i) for i in k)
    all_shape_total_table[kname] = summary_table[k]["total"]
    all_shape_mem_table[kname]   = summary_table[k]["mem_table"]


total_shape_tb     = pd.DataFrame(all_shape_total_table)
# 排序 index 排序 
total_shape_tb     = total_shape_tb.sort_index()
total_shape_mem_tb = pd.DataFrame(all_shape_mem_table)
# each shape has a table 

wb = Workbook()
ws = wb.active
ws.title = "total_shape_tb"
annotation = ["For lora finetune, the mem usage and time usage of ADAM can be ignored. "]
annotation += ["time(us)                activation mem (B)"]
annotation += ["total time and total activation mem usage"]
last_row = write_to_sheet(ws, annotation, 5, 5)
last_row = write_to_sheet(ws, total_shape_tb, last_row+3, 5)

annotation = ["TABLE of Each Layer Activation mem usage"]
last_row = write_to_sheet(ws, annotation, last_row+3, 5)
last_row = write_to_sheet(ws, total_shape_mem_tb, last_row+3, 5)


for k in summary_table.keys():
    kname = "_".join(str(i) for i in k)
    ws = wb.create_sheet(title=kname)
    annotation = ["         TABLE of Each Layer Time usage     "]
    last_row = write_to_sheet(ws, annotation, 5, 5)
    op_times_df = summary_table[k]["ops_times"]
    last_row = write_to_sheet(ws, op_times_df, last_row+3, 5)
    annotation = ["         TABLE of Detail Layey Information      "]
    last_row = write_to_sheet(ws, annotation, last_row+3, 5)
    fwd_bwd_df = summary_table[k]["fwd_bwd_df"]
    last_row = write_to_sheet(ws, fwd_bwd_df, last_row+2, 5)

wb.save("output.xlsx")
# total_shape_op_tb  = pd.DataFrame(all_shape_ops_table)


import ipdb;ipdb.set_trace()


# group by path_1, path_2, path_3, path_4

import pdb;pdb.set_trace()